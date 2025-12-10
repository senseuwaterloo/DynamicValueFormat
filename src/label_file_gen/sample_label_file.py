from utils.common import common_args

import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_random_percent(input_file, output_file, ratio=0.2):
    wb = load_workbook(input_file)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))

    k = int(len(rows) * ratio)

    selected_rows = random.sample(rows, k)

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in selected_rows:
        for cell in row:
            cell.fill = fill

    wb.save(output_file)

if __name__ == "__main__":
    args = common_args()
    dataset = args.dataset
    input_dir = args.input
    output_dir = args.output

    print(dataset + '\n')

    input_file = input_dir + '/' + dataset + '_labels.xlsx'
    output_file = output_dir + '/' + dataset + '_labels.xlsx'

    highlight_random_percent(input_file, output_file)